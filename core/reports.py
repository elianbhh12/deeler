"""Reportes: Excel consolidado del backlog (hoja única).

Paleta alineada a la identidad del banco (mismos colores que usa la interfaz,
ver core/config.py) — deliberadamente sobria: acento amarillo + negro para
encabezados, y verde/rojo solo puntuales para marcar estado, sin pintar filas
enteras (salvo el resaltado de HU ya desplegadas en PDN).
"""
from pathlib import Path
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import ICON_SUCCESS, ICON_FAIL, ACCENT, INK, GREEN, RED, MUTED
from .analysis import obtener_estado_pdn_real

#  Colores del reporte — mismos que usa la app (sin el "#"), no un set aparte.
_C_ACCENT = ACCENT.lstrip("#")   # amarillo banco — fondo de encabezado
_C_INK    = INK.lstrip("#")      # negro — texto de encabezado y bordes
_C_GREEN  = GREEN.lstrip("#")    # éxito
_C_RED    = RED.lstrip("#")      # error
_C_MUTED  = MUTED.lstrip("#")    # texto secundario / N/A
_C_LINE   = "9C9A98"             # borde — gris oscuro, visible sin ser negro puro
_C_BANDA  = "FAFAF9"             # banda de fila alterna, muy sutil (= SURFACE de la UI)

_C_PDN_BG = "D1FAE5"              # verde claro — mismo tono que usa la UI para "ok"

_FUENTE = "Calibri"               # una sola fuente en todo el archivo
_FONT_HEADER = Font(name=_FUENTE, bold=True, color=_C_INK)
_FONT_DATA   = Font(name=_FUENTE, color=_C_INK)
_ALIGN_CENTRO = Alignment(horizontal="center", vertical="center")
_BORDE = Border(
    left=Side(style="thin", color=_C_LINE),
    right=Side(style="thin", color=_C_LINE),
    top=Side(style="thin", color=_C_LINE),
    bottom=Side(style="thin", color=_C_LINE),
)


def _header_style(ws, fila=1):
    """Encabezado único y consistente: fondo amarillo banco, texto negro."""
    fill = PatternFill(start_color=_C_ACCENT, end_color=_C_ACCENT, fill_type="solid")
    for cell in ws[fila]:
        cell.fill = fill
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_CENTRO
        cell.border = _BORDE
    ws.row_dimensions[fila].height = 20
    ws.freeze_panes = f"A{fila + 1}"
    ws.auto_filter.ref = f"A{fila}:{get_column_letter(ws.max_column)}{fila}"


def _autofit_columnas(ws, minimo=6, maximo=45, holgura=3):
    """Ajusta el ancho de cada columna al contenido más largo que tenga
    (encabezado o dato), con un margen para el ícono de autofiltro y sin
    dejar que una celda gigante desborde toda la hoja."""
    for columna in ws.columns:
        letra = get_column_letter(columna[0].column)
        largo = max((len(str(c.value)) for c in columna if c.value is not None), default=0)
        ws.column_dimensions[letra].width = max(minimo, min(maximo, largo + holgura))


def _bandear_filas(ws, primera_fila_datos, ultima_fila):
    """Pinta filas alternas con una banda muy sutil para que se vea como una
    tabla real, sin tocar bordes/fuente/alineación (eso ya lo puso cada fila
    al construirse) y sin pisar celdas que ya tengan un color propio (los
    íconos ✓/✗ o el % de éxito)."""
    banda_fill = PatternFill(start_color=_C_BANDA, end_color=_C_BANDA, fill_type="solid")
    for fila in range(primera_fila_datos, ultima_fila + 1):
        if (fila - primera_fila_datos) % 2 == 0:
            continue
        for cell in ws[fila]:
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = banda_fill


def generar_excel_consolidado(resultados: list, guardar_en_carpeta: Path = None) -> bytes:
    """Genera el Excel Consolidado: ID/título/tipo/sprint/estado ADO, fechas
    de creación y cierre, y el despliegue real en PDN (resaltado en verde)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    #  HEADERS
    headers = ["ID", "Título", "Tipo", "Sprint", "Estado ADO",
               "Fecha Creación", "Fecha Cierre",
               "Desplegado en PDN por", "Fecha despliegue PDN"]

    ws.append(headers)
    _header_style(ws)

    #  DATOS
    row = 2
    for r in sorted(resultados, key=lambda x: x.get("downloaded_at", ""), reverse=True):
        hu_id = r.get("hu_id", "")
        title = r.get("hu_title", "")[:50]
        tipo = r.get("tipo_cambio", "")
        sprint = r.get("sprint", "?")
        estado_ado = r.get("estado_ado", "New")

        # Fechas reales de ADO
        created_date = r.get("created_date", "")
        changed_date = r.get("changed_date", "")

        fecha_creacion = created_date[:10] if created_date else "?"

        # Fecha de cierre: solo si está Closed
        if estado_ado == "Closed":
            fecha_cierre = changed_date[:10] if changed_date else "?"
        else:
            fecha_cierre = "-"  # Sin cerrar aún

        # Trazabilidad de despliegue real en PDN: se infiere de que TA/AID/UDZ
        # ya se subieron con éxito a la tabla PDN, no de un hecho manual (ver
        # core.analysis.obtener_estado_pdn_real).
        _pdn_real = obtener_estado_pdn_real(r)
        desplegado_pdn_por = _pdn_real["por"] or "-"
        desplegado_pdn_en_raw = _pdn_real["en"] or ""
        desplegado_pdn_en = desplegado_pdn_en_raw[:16].replace("T", " ") if desplegado_pdn_en_raw else "-"

        ws.append([
            hu_id, title, tipo, sprint, estado_ado,
            fecha_creacion, fecha_cierre,
            desplegado_pdn_por, desplegado_pdn_en
        ])

        # Aplicar estilos a fila: borde, centrado y una sola fuente en todo
        # el archivo — verde/rojo solo en los íconos ✓/✗, el resto en negro.
        # Una HU ya desplegada en PDN se resalta entera en verde claro — es
        # el hito que de verdad importa (la HU "termina" al llegar a PDN),
        # así se ve de un vistazo sin tener que leer columna por columna.
        _fill_pdn = PatternFill(start_color=_C_PDN_BG, end_color=_C_PDN_BG, fill_type="solid") if _pdn_real["desplegado"] else None
        for cell in ws[row]:
            cell.border = _BORDE
            cell.alignment = _ALIGN_CENTRO
            if ICON_SUCCESS in str(cell.value):
                cell.font = Font(name=_FUENTE, color=_C_GREEN)
            elif ICON_FAIL in str(cell.value):
                cell.font = Font(name=_FUENTE, color=_C_RED)
            else:
                cell.font = _FONT_DATA
            if _fill_pdn:
                cell.fill = _fill_pdn

        row += 1

    _bandear_filas(ws, 2, row - 1)

    #  AJUSTAR ANCHO COLUMNAS al contenido real (encabezado o dato más largo
    #  de cada columna), no a un valor fijo — así si cambia el texto de un
    #  encabezado o el largo típico de un dato, la celda se sigue viendo
    #  completa sin volver a tocar este número a mano.
    _autofit_columnas(ws)

    # Guardar a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Si se proporciona una carpeta, guardar también allí con nombre estándar
    if guardar_en_carpeta:
        guardar_en_carpeta.mkdir(parents=True, exist_ok=True)
        archivo_path = guardar_en_carpeta / "Consolidado_Backlog.xlsx"
        wb.save(str(archivo_path))

    return output.getvalue()
